import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import io
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 设置页面
st.set_page_config(page_title="LTV版本对比工具", layout="wide")

st.title("🎮 LTV版本对比分析工具")
st.markdown("上传两个版本的Global LTV数据，自动分析收益前N国家的LTV表现")

# ==================== 侧边栏配置 ====================
with st.sidebar:
    st.header("⚙️ 配置")
    
    # 选择LTV指标
    selected_ltvs = st.multiselect(
        "选择要分析的LTV指标",
        options=['LTV1', 'LTV7', 'LTV14', 'LTV30'],
        default=['LTV1', 'LTV7', 'LTV14']
    )
    
    # 选择Top N
    top_n = st.slider("显示Top N国家", min_value=3, max_value=10, value=5)
    
    st.markdown("---")
    st.markdown("""
    **📌 数据格式要求：**
    - CSV文件
    - BI报表上下载LTV数据，必须包含列：`weidu`(国家), `new_user`(新增用户)
    - LTV列：`ltv01`, `ltv07`, `ltv14` 等
    """)

# ==================== Excel美化函数 ====================
def apply_excel_style(filepath):
    """应用Excel样式美化"""
    wb = load_workbook(filepath)
    
    # ---------- 颜色定义 ----------
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    data_font = Font(name='微软雅黑', size=10)
    data_alignment_center = Alignment(horizontal='center', vertical='center')
    data_alignment_right = Alignment(horizontal='right', vertical='center')
    data_alignment_left = Alignment(horizontal='left', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    positive_fill = PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid')
    negative_fill = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
    positive_font = Font(name='微软雅黑', size=10, color='1A7A3A', bold=True)
    negative_font = Font(name='微软雅黑', size=10, color='C0392B', bold=True)
    
    even_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    odd_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    country_font = Font(name='微软雅黑', size=10, bold=True)
    
    # ---------- 处理每个Sheet ----------
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val:
                headers.append(val)
        
        change_cols = []
        country_col = None
        for idx, header in enumerate(headers, 1):
            if header and '变化%' in str(header):
                change_cols.append(idx)
            if header == '国家':
                country_col = idx
        
        # 表头
        ws.row_dimensions[1].height = 30
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 数据行
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 22
            row_fill = even_fill if (row_idx - 2) % 2 == 1 else odd_fill
            
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.fill = row_fill
                cell.border = thin_border
                
                if country_col and col_idx == country_col:
                    cell.font = country_font
                    cell.alignment = data_alignment_left
                    continue
                
                if col_idx in change_cols:
                    if isinstance(cell.value, (int, float)):
                        if cell.value > 0:
                            cell.font = positive_font
                            cell.fill = positive_fill
                            cell.value = f"+{cell.value:.2f}%"
                        elif cell.value < 0:
                            cell.font = negative_font
                            cell.fill = negative_fill
                            cell.value = f"{cell.value:.2f}%"
                        else:
                            cell.value = f"{cell.value:.2f}%"
                        cell.alignment = data_alignment_center
                    continue
                
                if isinstance(cell.value, (int, float)):
                    cell.alignment = data_alignment_right
                    if '用户' in str(headers[col_idx-1]) and cell.value >= 1000:
                        cell.value = f"{int(cell.value):,}"
                    elif 'LTV' in str(headers[col_idx-1]) or 'ltv' in str(headers[col_idx-1]).lower():
                        cell.value = round(cell.value, 4)
                else:
                    cell.alignment = data_alignment_center
        
        # 列宽
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    try:
                        val_len = len(str(cell.value))
                        if val_len > max_length:
                            max_length = val_len
                    except:
                        pass
            adjusted_width = min(max(max_length + 4, 14), 35)
            ws.column_dimensions[col_letter].width = adjusted_width
    
    wb.save(filepath)
    return filepath

# ==================== 核心函数 ====================
def load_and_clean(file):
    """加载并清洗CSV数据"""
    try:
        df = pd.read_csv(file)
    except UnicodeDecodeError:
        try:
            df = pd.read_csv(file, encoding='utf-16', sep='\t')
        except:
            df = pd.read_csv(file, encoding='latin-1')
    
    df = df.replace('', np.nan)
    df = df.replace(' ', np.nan)
    
    numeric_cols = ['new_user', 'rev00', 'dav00', 'ltv00', 'rev01', 'dav01', 'ltv01', 
                    'rev03', 'dav03', 'ltv03', 'rev07', 'dav07', 'ltv07', 'rev14', 'dav14', 'ltv14',
                    'rev30', 'dav30', 'ltv30']
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    return df

def get_top_countries(df, n=5):
    """计算收益前N国家"""
    if 'rev00' in df.columns:
        country_earnings = df.groupby('weidu')['rev00'].sum().sort_values(ascending=False)
    elif 'ltv01' in df.columns and 'new_user' in df.columns:
        df['_revenue_est'] = df['ltv01'] * df['new_user']
        country_earnings = df.groupby('weidu')['_revenue_est'].sum().sort_values(ascending=False)
    else:
        return []
    
    return country_earnings.head(n).index.tolist()

def calculate_country_metrics(df, country):
    """计算单个国家的LTV指标"""
    country_df = df[df['weidu'] == country]
    if len(country_df) == 0:
        return None
    
    metrics = {
        '国家': country,
        '新增用户': country_df['new_user'].sum() if 'new_user' in country_df.columns else 0,
    }
    
    ltv_cols = [col for col in ['ltv01', 'ltv07', 'ltv14', 'ltv30'] if col in df.columns]
    for col in ltv_cols:
        metrics[col] = country_df[col].mean()
    
    return metrics

def create_ltv_comparison(df_a, df_b, countries, version_a, version_b, ltv_cols):
    """创建LTV对比数据"""
    results = []
    
    for country in countries:
        metrics_a = calculate_country_metrics(df_a, country)
        metrics_b = calculate_country_metrics(df_b, country)
        
        if metrics_a is None or metrics_b is None:
            continue
        
        row = {
            '国家': country,
            f'{version_a}_用户': metrics_a['新增用户'],
            f'{version_b}_用户': metrics_b['新增用户'],
            '用户变化%': ((metrics_b['新增用户'] / metrics_a['新增用户'] - 1) * 100) if metrics_a['新增用户'] > 0 else 0,
        }
        
        for ltv_col in ltv_cols:
            val_a = metrics_a.get(ltv_col, 0)
            val_b = metrics_b.get(ltv_col, 0)
            row[f'{version_a}_{ltv_col}'] = val_a
            row[f'{version_b}_{ltv_col}'] = val_b
            row[f'{ltv_col}_变化%'] = ((val_b / val_a - 1) * 100) if val_a > 0 else 0
        
        results.append(row)
    
    return pd.DataFrame(results)

def create_global_summary(df_a, df_b, version_a, version_b, ltv_cols):
    """创建全球汇总"""
    summary = {'维度': ['Global']}
    
    for df, ver in [(df_a, version_a), (df_b, version_b)]:
        summary[f'{ver}_用户'] = df['new_user'].sum() if 'new_user' in df.columns else 0
        for col in ltv_cols:
            if col in df.columns:
                summary[f'{ver}_{col}'] = df[col].mean()
    
    for col in ltv_cols:
        val_a = summary.get(f'{version_a}_{col}', 0)
        val_b = summary.get(f'{version_b}_{col}', 0)
        summary[f'{col}_变化%'] = ((val_b / val_a - 1) * 100) if val_a > 0 else 0
    
    return pd.DataFrame([summary])

def create_ltv_bar_chart(df_compare, version_a, version_b, ltv_cols, title):
    """使用Plotly创建LTV对比柱状图"""
    countries = df_compare['国家'].tolist()
    n_metrics = len(ltv_cols)
    
    if n_metrics == 0:
        return None
    
    fig = make_subplots(rows=1, cols=n_metrics, 
                        subplot_titles=[f'{col.upper()}' for col in ltv_cols],
                        shared_yaxes=False)
    
    colors_a = '#2E86AB'
    colors_b = '#E84855'
    
    for idx, ltv_col in enumerate(ltv_cols, 1):
        col_a = f'{version_a}_{ltv_col}'
        col_b = f'{version_b}_{ltv_col}'
        
        if col_a not in df_compare.columns or col_b not in df_compare.columns:
            continue
        
        fig.add_trace(
            go.Bar(name=version_a, x=countries, y=df_compare[col_a],
                   marker_color=colors_a, text=df_compare[col_a].round(4),
                   textposition='outside', textfont=dict(size=10)),
            row=1, col=idx
        )
        
        fig.add_trace(
            go.Bar(name=version_b, x=countries, y=df_compare[col_b],
                   marker_color=colors_b, text=df_compare[col_b].round(4),
                   textposition='outside', textfont=dict(size=10)),
            row=1, col=idx
        )
        
        fig.update_xaxes(title_text="Country", tickangle=45, row=1, col=idx)
        fig.update_yaxes(title_text=ltv_col.upper(), row=1, col=idx)
    
    fig.update_layout(
        title_text=title,
        height=500,
        showlegend=True,
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        bargap=0.15
    )
    
    return fig

def create_daily_trend_chart(df_a, df_b, country, version_a, version_b, ltv_cols):
    """使用Plotly创建每日趋势图"""
    if len(ltv_cols) == 0:
        return None
    
    fig = make_subplots(rows=1, cols=len(ltv_cols), 
                        subplot_titles=[f'{col.upper()} Trend' for col in ltv_cols],
                        shared_yaxes=False)
    
    country_df_a = df_a[df_a['weidu'] == country]
    country_df_b = df_b[df_b['weidu'] == country]
    
    colors_a = '#2E86AB'
    colors_b = '#E84855'
    
    for idx, ltv_col in enumerate(ltv_cols, 1):
        if 'first_open_date_day' in country_df_a.columns and ltv_col in country_df_a.columns:
            daily_a = country_df_a.groupby('first_open_date_day')[ltv_col].mean().reset_index()
            if len(daily_a) > 0:
                fig.add_trace(
                    go.Scatter(x=daily_a['first_open_date_day'], y=daily_a[ltv_col],
                              mode='lines+markers', name=version_a,
                              marker=dict(size=8, color=colors_a),
                              line=dict(width=2, color=colors_a)),
                    row=1, col=idx
                )
        
        if 'first_open_date_day' in country_df_b.columns and ltv_col in country_df_b.columns:
            daily_b = country_df_b.groupby('first_open_date_day')[ltv_col].mean().reset_index()
            if len(daily_b) > 0:
                fig.add_trace(
                    go.Scatter(x=daily_b['first_open_date_day'], y=daily_b[ltv_col],
                              mode='lines+markers', name=version_b,
                              marker=dict(size=8, color=colors_b),
                              line=dict(width=2, color=colors_b)),
                    row=1, col=idx
                )
        
        fig.update_xaxes(title_text="Date", row=1, col=idx)
        fig.update_yaxes(title_text=ltv_col.upper(), row=1, col=idx)
    
    fig.update_layout(
        title_text=f'{country} - Daily LTV Trend',
        height=450,
        showlegend=True,
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
    )
    
    return fig

# ==================== 主界面 ====================
col1, col2 = st.columns(2)

with col1:
    st.subheader("📁 版本A (旧版本)")
    file_a = st.file_uploader("上传版本A的CSV文件", type=['csv', 'txt'], key='file_a')
    version_a = st.text_input("版本A名称", value="v1.0.0", key='ver_a')

with col2:
    st.subheader("📁 版本B (新版本)")
    file_b = st.file_uploader("上传版本B的CSV文件", type=['csv', 'txt'], key='file_b')
    version_b = st.text_input("版本B名称", value="v2.0.0", key='ver_b')

# 分析按钮
if st.button("🚀 开始分析", type="primary", use_container_width=True):
    if file_a is None or file_b is None:
        st.error("❌ 请上传两个版本的CSV文件")
        st.stop()
    
    with st.spinner("🔄 正在分析数据..."):
        try:
            # 加载数据
            df_a = load_and_clean(file_a)
            df_b = load_and_clean(file_b)
            
            # 获取LTV列
            ltv_cols = [col for col in ['ltv01', 'ltv07', 'ltv14', 'ltv30'] 
                       if col in df_a.columns and col in df_b.columns]
            
            if not ltv_cols:
                st.error("❌ 数据中找不到LTV列 (ltv01, ltv07, ltv14, ltv30)")
                st.stop()
            
            # 获取Top N国家
            countries = get_top_countries(df_a, top_n)
            if not countries:
                st.error("❌ 无法计算收益排名，请确认数据包含 'rev00' 或 'ltv01' 列")
                st.stop()
            
            # 创建对比数据
            df_compare = create_ltv_comparison(df_a, df_b, countries, version_a, version_b, ltv_cols)
            df_global = create_global_summary(df_a, df_b, version_a, version_b, ltv_cols)
            
            st.success(f"✅ 分析完成！Top {top_n} 国家: {', '.join(countries)}")
            
            # Tab显示
            tab1, tab2, tab3, tab4 = st.tabs(["📊 国家LTV对比", "🌍 全球汇总", "📈 趋势图表", "📥 下载报告"])
            
            with tab1:
                st.subheader(f"📊 Top {top_n} 国家 LTV 对比")
                
                display_df = df_compare.copy()
                for col in display_df.columns:
                    if '变化%' in col:
                        display_df[col] = display_df[col].apply(lambda x: f"{x:+.2f}%")
                    elif '用户' in col and '变化' not in col:
                        display_df[col] = display_df[col].apply(lambda x: f"{int(x):,}")
                    elif 'ltv' in col.lower():
                        display_df[col] = display_df[col].apply(lambda x: f"{x:.4f}")
                
                st.dataframe(display_df, use_container_width=True)
                
                fig = create_ltv_bar_chart(df_compare, version_a, version_b, ltv_cols, 
                                          f"Top {top_n} Countries LTV Comparison")
                if fig:
                    st.plotly_chart(fig, use_container_width=True)
            
            with tab2:
                st.subheader("🌍 全球汇总")
                display_global = df_global.copy()
                for col in display_global.columns:
                    if '变化%' in col:
                        display_global[col] = display_global[col].apply(lambda x: f"{x:+.2f}%")
                    elif 'ltv' in col.lower():
                        display_global[col] = display_global[col].apply(lambda x: f"{x:.4f}")
                    elif '用户' in col and '变化' not in col:
                        display_global[col] = display_global[col].apply(lambda x: f"{int(x):,}")
                st.dataframe(display_global, use_container_width=True)
            
            with tab3:
                st.subheader("📈 各国每日LTV趋势")
                
                if len(countries) > 0:
                    country_select = st.selectbox("选择国家查看趋势", countries)
                    
                    if country_select:
                        fig = create_daily_trend_chart(df_a, df_b, country_select, 
                                                       version_a, version_b, ltv_cols)
                        if fig:
                            st.plotly_chart(fig, use_container_width=True)
            
            with tab4:
                st.subheader("📥 下载报告")
                
                excel_buffer = io.BytesIO()
                game_name = "Game"
                
                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                    df_global.to_excel(writer, sheet_name='全球汇总', index=False)
                    df_compare.to_excel(writer, sheet_name='国家LTV对比', index=False)
                    
                    for country in countries:
                        country_df_a = df_a[df_a['weidu'] == country]
                        country_df_b = df_b[df_b['weidu'] == country]
                        
                        daily_a = pd.DataFrame()
                        daily_b = pd.DataFrame()
                        
                        if len(country_df_a) > 0 and 'first_open_date_day' in country_df_a.columns:
                            agg_dict = {col: 'mean' for col in ltv_cols if col in country_df_a.columns}
                            if 'new_user' in country_df_a.columns:
                                agg_dict['new_user'] = 'sum'
                            daily_a = country_df_a.groupby('first_open_date_day').agg(agg_dict).reset_index()
                            if not daily_a.empty:
                                daily_a['版本'] = version_a
                        
                        if len(country_df_b) > 0 and 'first_open_date_day' in country_df_b.columns:
                            agg_dict = {col: 'mean' for col in ltv_cols if col in country_df_b.columns}
                            if 'new_user' in country_df_b.columns:
                                agg_dict['new_user'] = 'sum'
                            daily_b = country_df_b.groupby('first_open_date_day').agg(agg_dict).reset_index()
                            if not daily_b.empty:
                                daily_b['版本'] = version_b
                        
                        if not daily_a.empty or not daily_b.empty:
                            daily_combined = pd.concat([daily_a, daily_b], ignore_index=True)
                            sheet_name = re.sub(r'[\\/*?:"<>|]', '', country)[:31]
                            daily_combined.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 应用样式美化
                excel_buffer.seek(0)
                temp_path = "/tmp/temp_lvt_report.xlsx"
                with open(temp_path, "wb") as f:
                    f.write(excel_buffer.getvalue())
                
                # 美化Excel
                apply_excel_style(temp_path)
                
                # 读取美化后的文件
                with open(temp_path, "rb") as f:
                    styled_buffer = io.BytesIO(f.read())
                
                st.download_button(
                    label="📥 下载",
                    data=styled_buffer,
                    file_name=f"{game_name}_LTV_Comparison.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
        
        except Exception as e:
            st.error(f"❌ 分析失败: {str(e)}")
            st.exception(e)

st.markdown("---")
st.caption("💡 上传两个版本的Global LTV数据，系统自动识别Top国家并进行对比分析")
